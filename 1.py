import os.path
from pypdf import PdfReader
from openpyxl import load_workbook
from xlrd import open_workbook

from zipfile import ZipFile
import PyPDF2
from io import BytesIO




# CURRENT_FILE = os.path.abspath(__file__)
# print(CURRENT_FILE)
#
# CURRENT_DIR = os.path.dirname(CURRENT_FILE)
# print(CURRENT_DIR)
#
# TMP_DIR = os.path.join(CURRENT_DIR, 'tmp')
# print(TMP_DIR)


# pdf_path = os.path.abspath('File_pdf.pdf')
# print(pdf_path)

# Работа с PDF
reader = PdfReader('File_folder/File_pdf.pdf')
print(reader.pages)
print(len(reader.pages))
print(reader.pages[1])
print(reader.pages[1].extract_text())
assert "Теория физически корректного рендеринга и затенения" in reader.pages[1].extract_text()
print(os.path.getsize('File_folder/File_pdf.pdf'))
assert os.path.getsize('File_folder/File_pdf.pdf') == 1164287


# Работа с XLSX
workbook = load_workbook('File_folder/File_xls.xlsx')
sheet = workbook.active
print(sheet.cell(row=1, column=2).value)

print("Содержимое всего листа построчно (кортежи):")
for row in sheet.iter_rows(values_only=True):
    print(row)

print("Содержимое всего листа одной строкой (список кортежей):")
print(list(sheet.iter_rows(values_only=True)))

# Работа с XLS
workbook = open_workbook('File_folder/File.xls')
print(workbook.nsheets)
print(workbook.sheet_names())
sheet = workbook.sheet_by_index(0)
print(sheet.nrows)
print(sheet.ncols)
print(sheet.cell_value(rowx=1,colx=1))

for rx in range(sheet.nrows):
    print(sheet.row(rx))


# Работа с архивами (если в архиве один файл?)


with ZipFile('resources/documents.zip') as zip_file:
    print(zip_file.namelist())
    text = zip_file.read('вставить имя файла')
    print(text)
    # Извлечение архива
    zip_file.extract('вставить имя архива')







